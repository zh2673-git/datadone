#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Excel导出器 - 生成精简的Excel分析报告

精简后10个工作表（原10个）：
  合并：行为基线→风险研判表（风险研判自带基准）
  合并：行为模式+时序链→行为发现表（同层级发现归一）
  新增：存取现明细表（逐笔存取现+大额标记+备注）
  新增：重点收支明细表（重点收支+特殊金额+特殊日期+规避阈值）
  优化：综合分析表增加交叉类型列
  保留：汇总表、账单频率、话单频率、异常明细、大额资金跟踪
"""

import logging
import os
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models.report import ReportData, ExportConfig

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel报告导出器 - 精简版"""

    # 工作表1：分析汇总表列
    SUMMARY_COLUMNS = ['分析类型', '平台', '数据来源', '本方姓名', '存现金额', '取现金额', '转入金额', '转出金额']

    # 工作表2：存取现明细表列（新增）
    CASH_DETAIL_COLUMNS = [
        '本方姓名', '交易日期', '存取现标识', '交易金额', '对方姓名',
        '银行类型', '账户余额', '大额标记', '备注'
    ]

    # 工作表3：重点收支明细表列（新增）
    KEY_TRANSACTION_COLUMNS = [
        '本方姓名', '交易日期', '交易方向', '交易金额', '对方姓名',
        '重点类型', '重点子类', '特征描述', '置信度', '数据来源'
    ]

    # 工作表3：账单类频率表列
    BILL_FREQUENCY_COLUMNS = [
        '平台', '数据来源', '本方姓名', '对方姓名',
        '收入总额', '支出总额', '交易次数', '交易总金额', '收入占比', '支出占比'
    ]

    # 工作表4：话单类频率表列
    CALL_FREQUENCY_COLUMNS = [
        '平台', '数据来源', '本方姓名', '对方姓名', '对方号码',
        '对方单位名称', '对方职务', '通话次数', '主叫次数', '被叫次数',
        '通话总时长秒', '通话总时长分钟', '通话占比', '特殊时间次数'
    ]

    # 工作表5：综合分析表列（增加交叉类型）
    CROSS_ANALYSIS_COLUMNS = [
        '分析基准', '交叉类型', '本方姓名', '对方姓名', '对方号码', '对方单位名称', '对方职务',
        '通话次数', '通话总时长分钟', '收入总额', '支出总额', '交易次数',
        '平台', '平台金额分布',
        '银行_收入总额', '银行_支出总额', '银行_交易次数',
        '微信_收入总额', '微信_支出总额', '微信_交易次数',
        '支付宝_收入总额', '支付宝_支出总额', '支付宝_交易次数'
    ]

    # 工作表6：大额资金跟踪表列
    FUND_TRACKING_COLUMNS = [
        '分析类型', '追踪层级', '核心人员', '关联人员', '交易日期', '交易金额',
        '交易方向', '大额级别', '数据来源', '资金流向', '追踪说明',
        '月份', '通话日期', '通话对方', '通话对方单位'
    ]

    # 工作表7：异常明细表列
    ANOMALY_COLUMNS = [
        '本方姓名', '异常类型', '异常子类', '偏离度', '严重程度',
        '对方姓名', '交易日期', '交易金额', '基线均值', '说明'
    ]

    # 工作表8：行为发现表列（行为模式+时序链合并）
    BEHAVIOR_FINDING_COLUMNS = [
        '本方姓名', '发现类型', '编号', '名称', '匹配度/强度',
        '涉及对手', '关键证据', '附加信息', '报告用语'
    ]

    # 工作表9：风险研判表列（含基线指标）
    # 基线指标在前，风险研判在后
    BASELINE_COLUMNS_FOR_RISK = [
        '数据充足度', '数据月数',
        '月均收入', '月均收入_标准差', '月均支出', '月均支出_标准差',
        '月均交易次数', '月均交易对手数',
        '单笔金额均值', '单笔金额中位数', '单笔金额P25', '单笔金额P75',
        '工作时间交易占比', '深夜交易占比', '周末交易占比',
        '存取现月均金额', '存取现占收支比',
        '月均通话次数', '月均通话时长分钟',
        '收入趋势', '支出趋势', '对手数趋势',
    ]

    RISK_COLUMNS_FOR_RISK = [
        '综合风险分数', '综合风险等级',
        '异常偏离度得分', '异常偏离度说明',
        '行为模式得分', '行为模式说明',
        '证据链得分', '证据链说明',
        '规模得分', '规模说明',
        '证据充分度', '已有证据', '待补充证据',
        '重点人员', '调查方向建议', '重点时段'
    ]

    RISK_WITH_BASELINE_COLUMNS = ['本方姓名'] + BASELINE_COLUMNS_FOR_RISK + RISK_COLUMNS_FOR_RISK

    HABIT_INTEREST_COLUMNS = [
        '本方姓名', '类别', '子类', '证据类型', '匹配关键词',
        '交易次数', '总金额', '首次交易日期', '最近交易日期',
        '月均频次', '习惯等级', '典型时段', '代表性交易'
    ]

    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)

    def export(self, report_data: ReportData, config: Optional[ExportConfig] = None) -> str:
        if config is None:
            config = ExportConfig()

        output_dir = config.output_dir
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "分析报告.xlsx")

        self.logger.info("开始生成Excel报告...")

        wb = Workbook()
        wb.remove(wb.active)

        try:
            # 工作表1：分析汇总表
            self._write_summary_table(wb, report_data)

            # 工作表2：存取现明细表（新增）
            self._write_cash_detail(wb, report_data)

            # 工作表3：重点收支明细表（新增）
            self._write_key_transaction(wb, report_data)

            # 工作表4：账单类频率表
            self._write_bill_frequency(wb, report_data)

            # 工作表5：话单类频率表
            self._write_call_frequency(wb, report_data)

            # 工作表6：综合分析表（增强：交叉类型）
            self._write_cross_analysis(wb, report_data)

            # 工作表7：大额资金跟踪表
            self._write_fund_tracking(wb, report_data)

            # 工作表8：异常明细表
            self._write_anomalies(wb, report_data)

            # 工作表9：行为发现表（行为模式+时序链合并）
            self._write_behavior_finding(wb, report_data)

            # 工作表10：风险研判表（含基线指标）
            self._write_risk_with_baseline(wb, report_data)

            # 工作表11：习惯兴趣表（新增）
            self._write_habit_interest(wb, report_data)

            # 应用条件格式
            self._apply_conditional_formatting(wb, config)

            wb.save(output_path)
            self.logger.info(f"Excel报告已生成: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Excel导出失败: {e}", exc_info=True)
            raise

    # ------------------------------------------------------------------ #
    #  通用写入方法
    # ------------------------------------------------------------------ #

    def _write_sheet_from_dicts(self, wb: Workbook, sheet_name: str,
                                 columns: list, data: list) -> None:
        """通用方法：将dict列表写入工作表"""
        ws = wb.create_sheet(title=sheet_name)

        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, '')
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    value = ''
                elif hasattr(value, 'isoformat'):
                    value = value.isoformat()
                elif isinstance(value, (pd.Timestamp,)):
                    value = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        self._auto_column_width(ws, columns)
        ws.freeze_panes = 'A2'
        if data:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(data) + 1}"

    def _auto_column_width(self, ws, columns: list) -> None:
        """自动调整列宽"""
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name)) * 2
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=min(ws.max_row, 100)):
                for cell in row:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        try:
                            cn_count = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                            cell_len = cell_len + cn_count
                        except Exception:
                            pass
                        max_length = max(max_length, min(cell_len, 50))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    # ------------------------------------------------------------------ #
    #  各工作表写入
    # ------------------------------------------------------------------ #

    def _write_summary_table(self, wb: Workbook, report_data: ReportData) -> None:
        self._write_sheet_from_dicts(wb, '分析汇总表', self.SUMMARY_COLUMNS, report_data.summary_table)

    def _write_cash_detail(self, wb: Workbook, report_data: ReportData) -> None:
        """工作表2：存取现明细表（新增）"""
        self._write_sheet_from_dicts(wb, '存取现明细表', self.CASH_DETAIL_COLUMNS, report_data.cash_detail_data)

    def _write_key_transaction(self, wb: Workbook, report_data: ReportData) -> None:
        """工作表3：重点收支明细表（新增）"""
        self._write_sheet_from_dicts(wb, '重点收支明细表', self.KEY_TRANSACTION_COLUMNS, report_data.key_transaction_data)

    def _write_bill_frequency(self, wb: Workbook, report_data: ReportData) -> None:
        self._write_sheet_from_dicts(wb, '账单类频率表', self.BILL_FREQUENCY_COLUMNS, report_data.bill_frequency)

    def _write_call_frequency(self, wb: Workbook, report_data: ReportData) -> None:
        mapped_data = []
        for row in report_data.call_frequency:
            mapped = {col: row.get(col, '') for col in self.CALL_FREQUENCY_COLUMNS}
            mapped_data.append(mapped)
        self._write_sheet_from_dicts(wb, '话单类频率表', self.CALL_FREQUENCY_COLUMNS, mapped_data)

    def _write_cross_analysis(self, wb: Workbook, report_data: ReportData) -> None:
        """综合分析表（增强：交叉类型列+排序）"""
        all_rows = []
        for base_name, rows in report_data.cross_analysis.items():
            for row in rows:
                row_with_base = dict(row)
                row_with_base['分析基准'] = base_name

                # 计算交叉类型
                call_count = row.get('通话次数')
                income = row.get('收入总额')
                expense = row.get('支出总额')

                has_call = call_count is not None and not (isinstance(call_count, float) and pd.isna(call_count)) and call_count > 0
                has_fund = ((income is not None and not (isinstance(income, float) and pd.isna(income)) and income > 0) or
                            (expense is not None and not (isinstance(expense, float) and pd.isna(expense)) and expense > 0))

                if has_call and has_fund:
                    row_with_base['交叉类型'] = '双平台交叉'
                elif has_call:
                    row_with_base['交叉类型'] = '仅话单'
                else:
                    row_with_base['交叉类型'] = '仅账单'

                all_rows.append(row_with_base)

        # 排序：双平台交叉优先
        order = {'双平台交叉': 0, '仅账单': 1, '仅话单': 2}
        all_rows.sort(key=lambda x: (x.get('本方姓名', ''), order.get(x.get('交叉类型', ''), 9)))

        self._write_sheet_from_dicts(wb, '综合分析表', self.CROSS_ANALYSIS_COLUMNS, all_rows)

    def _write_fund_tracking(self, wb: Workbook, report_data: ReportData) -> None:
        self._write_sheet_from_dicts(wb, '大额资金跟踪表', self.FUND_TRACKING_COLUMNS, report_data.fund_tracking)

    def _write_anomalies(self, wb: Workbook, report_data: ReportData) -> None:
        """工作表7：异常明细表"""
        self._write_sheet_from_dicts(wb, '异常明细表', self.ANOMALY_COLUMNS, report_data.anomaly_data)

    def _write_behavior_finding(self, wb: Workbook, report_data: ReportData) -> None:
        """工作表8：行为发现表（行为模式+时序链合并）"""
        findings = []

        # 行为模式
        for p in report_data.pattern_data:
            additional = []
            if p.get('满足条件'):
                additional.append(f"满足：{p['满足条件']}")
            if p.get('未满足条件'):
                additional.append(f"未满足：{p['未满足条件']}")

            findings.append({
                '本方姓名': p.get('本方姓名', ''),
                '发现类型': '行为模式',
                '编号': p.get('模式编号', ''),
                '名称': p.get('模式名称', ''),
                '匹配度/强度': p.get('匹配度', ''),
                '涉及对手': p.get('涉及对手', ''),
                '关键证据': p.get('关键证据', ''),
                '附加信息': '；'.join(additional) if additional else '',
                '报告用语': p.get('报告用语', ''),
            })

        # 时序链
        for t in report_data.timeline_data:
            additional_parts = []
            if t.get('重复次数'):
                additional_parts.append(f"重复{t['重复次数']}次")
            if t.get('时间窗口小时'):
                additional_parts.append(f"时间窗口{t['时间窗口小时']}小时")

            findings.append({
                '本方姓名': t.get('本方姓名', ''),
                '发现类型': '时序链',
                '编号': '',
                '名称': t.get('链模式', ''),
                '匹配度/强度': t.get('链强度', ''),
                '涉及对手': t.get('对方姓名', ''),
                '关键证据': t.get('关键证据描述', ''),
                '附加信息': '；'.join(additional_parts) if additional_parts else '',
                '报告用语': t.get('事件序列', ''),
            })

        # 排序：行为模式在前，时序链在后；同类型按匹配度/强度降序
        type_order = {'行为模式': 0, '时序链': 1}
        findings.sort(key=lambda x: (x.get('本方姓名', ''), type_order.get(x.get('发现类型', ''), 9)))

        self._write_sheet_from_dicts(wb, '行为发现表', self.BEHAVIOR_FINDING_COLUMNS, findings)

    def _write_risk_with_baseline(self, wb: Workbook, report_data: ReportData) -> None:
        """工作表9：风险研判表（含基线指标）"""
        # 构建基线索引
        baseline_by_person = {}
        for bl in report_data.baseline_data:
            person = bl.get('本方姓名', '')
            if person:
                baseline_by_person[person] = bl

        # 合并基线+风险
        merged_data = []
        for risk_row in report_data.risk_data:
            person = risk_row.get('本方姓名', '')
            bl = baseline_by_person.get(person, {})

            merged = {'本方姓名': person}

            # 基线指标
            for col in self.BASELINE_COLUMNS_FOR_RISK:
                val = bl.get(col, '')
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    val = ''
                merged[col] = val

            # 风险研判指标
            for col in self.RISK_COLUMNS_FOR_RISK:
                val = risk_row.get(col, '')
                if val is None or (isinstance(val, float) and pd.isna(val)):
                    val = ''
                merged[col] = val

            merged_data.append(merged)

        self._write_sheet_from_dicts(wb, '风险研判表', self.RISK_WITH_BASELINE_COLUMNS, merged_data)

    # ------------------------------------------------------------------ #
    #  条件格式
    # ------------------------------------------------------------------ #

    def _apply_conditional_formatting(self, wb: Workbook, config: ExportConfig) -> None:
        """条件格式：大额标红、存取现高亮、风险等级着色"""
        if not config.highlight_large_amount:
            return

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        threshold = config.large_amount_threshold

        for ws in wb.worksheets:
            # 大额金额标红
            amount_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header in ('交易金额', '转入金额', '转出金额', '存现金额', '取现金额'):
                    amount_col = col_idx
                    break

            if amount_col is not None:
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=amount_col).value
                    try:
                        if cell_value is not None and abs(float(cell_value)) >= threshold:
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = red_fill
                                ws.cell(row=row_idx, column=col_idx).font = red_font
                    except (ValueError, TypeError):
                        pass

            # 大额标记列着色（★行高亮）
            marker_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == '大额标记':
                    marker_col = col_idx
                    break

            if marker_col is not None:
                orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=marker_col).value
                    if cell_value and '★' in str(cell_value):
                        ws.cell(row=row_idx, column=marker_col).fill = orange_fill

            # 存取现标识高亮
            cash_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == '存取现标识':
                    cash_col = col_idx
                    break

            if cash_col is not None:
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=cash_col).value
                    if cell_value in ('存现', '取现'):
                        ws.cell(row=row_idx, column=cash_col).fill = yellow_fill

            # 风险等级着色
            risk_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == '综合风险等级':
                    risk_col = col_idx
                    break

            if risk_col is not None:
                risk_colors = {
                    "高风险": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "中风险": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                    "低风险": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                }
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=risk_col).value
                    if cell_value in risk_colors:
                        ws.cell(row=row_idx, column=risk_col).fill = risk_colors[cell_value]

            # 严重程度着色
            severity_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == '严重程度':
                    severity_col = col_idx
                    break

            if severity_col is not None:
                severity_colors = {
                    "高": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                }
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=severity_col).value
                    if cell_value in severity_colors:
                        ws.cell(row=row_idx, column=severity_col).fill = severity_colors[cell_value]

            # 交叉类型着色
            cross_type_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == '交叉类型':
                    cross_type_col = col_idx
                    break

            if cross_type_col is not None:
                cross_colors = {
                    "双平台交叉": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                }
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=cross_type_col).value
                    if cell_value in cross_colors:
                        ws.cell(row=row_idx, column=cross_type_col).fill = cross_colors[cell_value]

            # 重点类型着色（重点收支明细表）
            key_type_col = None
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == '重点类型':
                    key_type_col = col_idx
                    break

            if key_type_col is not None:
                key_type_colors = {
                    "规避阈值": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    "特殊金额": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
                    "特殊日期": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
                    "工作收入": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                }
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=key_type_col).value
                    if cell_value in key_type_colors:
                        ws.cell(row=row_idx, column=key_type_col).fill = key_type_colors[cell_value]

    def _write_habit_interest(self, wb: Workbook, report_data: ReportData) -> None:
        """工作表11：习惯兴趣表"""
        self._write_sheet_from_dicts(
            wb, '习惯兴趣表', self.HABIT_INTEREST_COLUMNS, report_data.habit_interest_data
        )

        habit_colors = {
            "高频习惯": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "低频偏好": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            "偶尔消费": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        }
        ws = wb['习惯兴趣表']
        habit_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == '习惯等级':
                habit_col = col_idx
                break
        if habit_col is not None:
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=habit_col).value
                if val in habit_colors:
                    ws.cell(row=row_idx, column=habit_col).fill = habit_colors[val]
